"""
    Author: Pashazera
    Time  : 2023/4/13 9:03
    File  : clean.py
    describe :清洗物业表格，让物业数据规范化
"""

import  openpyxl as op
#
# filename2 = r'C:\Users\pashazera\Desktop\一标三实\数据\wy数据（xlsx）\骧龙国际A12-22栋人口信息.xlsx'
# ex2 = op.load_workbook(filename=filename2,data_only=True)
# data = ex2["12栋"]
#
#
# for i in range(4,data.max_row):
#
#     if data.cell(row = i,column = 1).value is None and data.cell(row = i,column = 2).value is not None :
#         value = data.cell(row = i -1,column = 1).value
#         data.cell(i , 1,value)
#     if data.cell(row = i,column = 1).value is None and data .cell(row = i,column = 2).value is None :
#         data.delete_rows(i)
#
# for i in range(2,data.max_row):
#     print(data.cell(row = i,column = 1).value)
#
# ex.save(filename)
# ex.save(r'C:\Users\pashazera\Desktop\data.xlsx')
# print(data.max_row)
# value = data.cell(row = 4,column = 2).value
# print(value)
# print(data.max_row)
# print(ex2.sheetnames)
#
# print(data.cell(3,4).value)

wb = op.Workbook()
ws = wb.create_sheet("骧龙国际A区",0)
ws_total = 2

# 创建表行

ws.cell(1 , 1 , "id")
ws.cell(1 , 2 ,"house_address")
ws.cell(1,3,"name")
ws.cell(1,4,"age")
ws.cell(1,5,'sex')
ws.cell(1,6,"person_class")
ws.cell(1,7,'id_card')
ws.cell(1,8,"phone_number")
ws.cell(1,9,"record")
ws.cell(1,10,'remark')


filename1 = r'C:\Users\pashazera\Desktop\一标三实\数据\wy数据（xlsx）\骧A1-11人员信息（2023.3.31）(1).xlsx'
ex1 = op.load_workbook(filename=filename1,data_only=True)

sheet = ex1.sheetnames
data = ex1[sheet[0]]

for i in  range(3,data.max_row):
    ws.cell(ws_total,1,ws_total - 1)
    ws.cell(ws_total,2,data.cell(i,1).value)
    ws.cell(ws_total,3,data.cell(i,2).value)
    ws.cell(ws_total,4,data.cell(i,4).value)
    ws.cell(ws_total,5,data.cell(i,3).value)
    ws.cell(ws_total,7,data.cell(i,6).value)
    ws.cell(ws_total,8,data.cell(i,5).value)
    ws.cell(ws_total,9,data.cell(i,7).value)
    ws.cell(ws_total,10,data.cell(i,8).value)
    ws_total = ws_total + 1
    if data.cell(i+1,1).value is None:
        break

filename2 = r"C:\Users\pashazera\Desktop\一标三实\数据\wy数据（xlsx）\骧龙国际A12-22栋人口信息.xlsx"
ex2 = op.load_workbook(filename=filename2,data_only=True)

sheet = ex2.sheetnames
for s in sheet:
    for i in range(2,ex2[s].max_row + 1):
      if ex2[s].cell(i,2).value is not None :
          ws.cell(ws_total, 1, ws_total - 1)
          ws.cell(ws_total, 2, ex2[s].cell(i, 2).value)
          ws.cell(ws_total, 3, ex2[s].cell(i, 3).value)
          ws.cell(ws_total, 4, ex2[s].cell(i, 7).value)
          if ex2[s].cell(i,6).value is not None:
              if int(ex2[s].cell(i, 6).value[16]) % 2 == 1:
                  ws.cell(ws_total, 5, '男')
              else:
                  ws.cell(ws_total, 5, '女')


          ws.cell(ws_total, 6, ex2[s].cell(i, 5).value)
          ws.cell(ws_total, 7, ex2[s].cell(i, 6).value)
          ws.cell(ws_total, 8, ex2[s].cell(i, 4).value)
          ws_total = ws_total + 1
      if ex2[s].cell(i,2).value is  None and ex2[s].cell(i,3).value is not None:
          ws.cell(ws_total, 1, ws_total - 1)
          ws.cell(ws_total, 2, ws.cell(ws_total-1,2).value)
          ws.cell(ws_total, 3, ex2[s].cell(i, 3).value)
          ws.cell(ws_total, 4, ex2[s].cell(i, 7).value)
          if ex2[s].cell(i, 6).value is not None:
              if int(ex2[s].cell(i, 6).value[16]) % 2 == 1:
                  ws.cell(ws_total, 5, '男')
              else:
                  ws.cell(ws_total, 5, '女')
          ws.cell(ws_total, 6, ex2[s].cell(i, 5).value)
          ws.cell(ws_total, 7, ex2[s].cell(i, 6).value)
          ws.cell(ws_total, 8, ex2[s].cell(i, 4).value)
          ws_total = ws_total + 1


wb.save("data.xlsx")











