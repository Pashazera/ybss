"""
    Author: Pashazera
    Time  : 2023/4/17 18:11
    File  : ygxsm.py
    describe :
"""

import openpyxl as op

filename = r'C:\Users\pashazera\Desktop\一标三实\数据\wy数据（xlsx）\营盘花苑疫苗接种名细.xlsx'
ex = op.load_workbook(filename=filename, data_only=True, read_only=True)
sheet = ex.sheetnames

wb = op.Workbook()
ws = wb.create_sheet('yphy', 0)
ws.cell(1, 1, 'houseaddress')
ws.cell(1, 2, 'type')
ws.cell(1, 3, 'name')
ws.cell(1, 4, 'age')
ws.cell(1, 5, 'card')
ws.cell(1, 6, 'phone')
ws.cell(1, 7, '1stitch')
ws.cell(1, 8, '2stitch')
ws.cell(1, 9, '3stitch')
ws.cell(1, 10, 'remark')

total = 2

for s in sheet:
    for i in range(3, ex[s].max_row + 1):
        if ex[s].cell(i, 1).value is not None:
            ws.cell(total, 1, ex[s].cell(i, 1).value)
            ws.cell(total, 2, ex[s].cell(i, 2).value)
            ws.cell(total, 3, ex[s].cell(i, 3).value)
            ws.cell(total, 4, ex[s].cell(i, 4).value)
            ws.cell(total, 5, ex[s].cell(i, 5).value)
            ws.cell(total, 6, ex[s].cell(i, 6).value)
            ws.cell(total, 7, ex[s].cell(i, 7).value)
            ws.cell(total, 8, ex[s].cell(i, 8).value)
            ws.cell(total, 9, ex[s].cell(i, 9).value)
            ws.cell(total, 10, ex[s].cell(i, 10).value)
            total = total + 1

        if ex[s].cell(i, 1).value is None and ex[s].cell(i, 2).value is None:
            ws.cell(total, 1, ws.cell(total - 1, 1).value)
            ws.cell(total, 2, ws.cell(total - 1, 2).value)
            ws.cell(total, 3, ex[s].cell(i, 3).value)
            ws.cell(total, 4, ex[s].cell(i, 4).value)
            ws.cell(total, 5, ex[s].cell(i, 5).value)
            ws.cell(total, 6, ex[s].cell(i, 6).value)
            ws.cell(total, 7, ex[s].cell(i, 7).value)
            ws.cell(total, 8, ex[s].cell(i, 8).value)
            ws.cell(total, 9, ex[s].cell(i, 9).value)
            ws.cell(total, 10, ex[s].cell(i, 10).value)
            total = total + 1

        if ex[s].cell(i, 1).value is None and ex[s].cell(i, 2).value is not None:
            ws.cell(total, 1, ws.cell(total - 1, 1).value)
            ws.cell(total, 2, ex[s].cell(i, 2).value)
            ws.cell(total, 3, ex[s].cell(i, 3).value)
            ws.cell(total, 4, ex[s].cell(i, 4).value)
            ws.cell(total, 5, ex[s].cell(i, 5).value)
            ws.cell(total, 6, ex[s].cell(i, 6).value)
            ws.cell(total, 7, ex[s].cell(i, 7).value)
            ws.cell(total, 8, ex[s].cell(i, 8).value)
            ws.cell(total, 9, ex[s].cell(i, 9).value)
            ws.cell(total, 10, ex[s].cell(i, 10).value)
            total = total + 1

        # if ex[s].cell(i, 1).value is None and ex[s].cell(i, 2).value is None and ex[s].cell(i, 3).value is None and i:
        #     break
        # total = total + 1

wb.save('yphy.xlsx')

print(sheet)
print(ex[sheet[0]].max_row)